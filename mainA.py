from PyQt5.QtWidgets import QMainWindow,QApplication,QFileDialog
from PyQt5 import QtWidgets
import bildirim
from vet_arayuz import Ui_MainWindow as vet_arayuz
import os
import sqlite3
import sys
import pandas
from openpyxl import load_workbook
import csv

class vet_sinifi(QMainWindow,vet_arayuz):
    def __init__(self):
        super().__init__()
        self.tablo_ = []
        self.columns = []
        self.tablo_adi = None
        self.imlec = None
        self.baglanti = None
        self.veri = None
        self.buttonN = None
        self.buttonY = None
        self.buttonO = None
        self.kutu = None
        self.uyari_kutu = None
        self.sutun_isimleri = []
        self.tutulan_uzanti = None
        self.ac_pencere()
            
    def dosya_sec(self):
        try:
            self.dosya_iletisim = QFileDialog(self)
            self.dosya_ac = self.dosya_iletisim.getOpenFileName(self,"Dosya Ac")
            return self.dosya_ac[0]
        except Exception as hata:
            self.statusbar.showMessage("Lutfen tekrar deneyin!Dosya seciminde bir hata ile karsilasildi:"+str(hata),5000)
    def temizle_tablo(self):
        #try:
        self.tablo.clear()
        #except Exception as hata:
         #   self.statusbar.showMessage("Lutfen uygulamayi yeniden acin!Bir hata ile karsilasildi:"+str(hata),5000)
    def sutun_ismi_yerlestir(self):
        self.imlec.execute(f"PRAGMA table_info({self.tablo_adi[0][0]})")
        gecis_sutun_adi = self.imlec.fetchall()
        for c in range(self.get_sutun_sayisi()):
            self.sutun_isimleri.insert(c,gecis_sutun_adi[c][1])
        self.tablo.setHorizontalHeaderLabels(self.sutun_isimleri)
        self.sutun_isimleri.clear()
    def guncel_tablo(self):
        #try:
        self.sutun_ismi_yerlestir()
        self.imlec.execute("SELECT * FROM {ta}".format(ta=self.tablo_adi[0][0]))
        self.veri = self.imlec.fetchall()
        self.tablo_ = [[str(j) for j in i] for i in self.veri]
        self.columns = [i[0] for i in self.imlec.description]
        for satir_say, satir_veri in enumerate(self.veri):
            for sutun_say, self.veri in enumerate(satir_veri):
                self.tablo.setItem(satir_say,sutun_say,QtWidgets.QTableWidgetItem(str(self.veri)))
        #except Exception as hata:
         #   self.statusbar.showMessage("Lutfen uygulamayi yeniden acin!Bir hata ile karsilasildi:"+str(hata),5000)
    def list_donusum_str(self,liste):
        metin = ""
        for i in liste:
            metin += i
        return metin
    def uzanti_bul(self,dosya):
        #bu dosya yolunun sonundaki dosyanin uzantisini ayiran dongu
        k = 0
        kntrl = False
        uzanti = []
        for i,parca in enumerate(dosya):
            if parca == '.' or kntrl == True:
                kntrl = True
                uzanti.insert(k,dosya[i])
                k += 1
        return uzanti
    def ayirici_tespit_et(self,dosya):
        """
        Bu fonksiyon dosyanin ilk satirina bakarak hangi ayiricinin 
        kullanildigini tepsit eder ve onu dondurur.
        """
        with open(dosya,'r',encoding="utf8") as d:
            bas_satir = d.readline()
            if bas_satir.find(";")!=-1:
                return ";"
            if bas_satir.find(",")!=-1:
                return ","
        return ";" #varsayilan ayırıcı
    def veri_yukle_sql(self,dosya):
        try:
            self.baglanti = sqlite3.connect(dosya)
            self.imlec = self.baglanti.cursor()
            self.imlec.execute("SELECT name FROM sqlite_master")
            self.tablo_adi = self.imlec.fetchall()
            if self.tablo_adi == []:
                bildirim.uyari_kutusu("Seçtiğiniz dosyada bir tablo ve veri bulunmaması sebebiyle yüklenemez.")
            else:
                self.temizle_tablo()
                self.guncel_tablo()
        except sqlite3.DatabaseError :
            bildirim.uyari_kutusu("Yüklemeye çalıştığınız dosya bir veritabanı değildir.")
        except Exception as hata:
            self.statusbar.showMessage("Lutfen uygulamayi yeniden acin!Bir yukleme hatasi ile karsilasildi:"+str(hata),5000)
    def veri_yukle_csv(self,dosya):
        try:
            self.temizle_tablo()
            self.veriler_csv = pandas.read_csv(dosya,delimiter=self.ayirici_tespit_et(dosya))
            satir = 0
            for sutun in range(0,len(self.veriler_csv.loc[0][:])):
                self.tablo.setItem(0,sutun,QtWidgets.QTableWidgetItem(str(self.veriler_csv.columns[sutun])))
            for satir_num in range(1,len(self.veriler_csv)+1):
                for sutun_num in range(0,len(self.veriler_csv.loc[0][:])):
                    self.tablo.setItem(satir_num,sutun_num,QtWidgets.QTableWidgetItem(str(self.veriler_csv.loc[satir][sutun_num])))
                satir += 1
        except pandas.errors.EmptyDataError:
            bildirim.uyari_kutusu("Girmeye çalıştığınız dosya boş gözüküyor.")
    def veri_yukle_xlsx(self,dosya):
        self.temizle_tablo()
        secili_d = load_workbook(dosya)
        veriler = secili_d.active
        for satir in range(1,veriler.max_row+1):
            for sutun in range(1,len(veriler[1])+1):
                self.tablo.setItem(satir-1,sutun-1,QtWidgets.QTableWidgetItem(str(veriler.cell(satir,sutun).value)))
    def veri_yukle(self):
        #try:
        dosya = self.dosya_sec()
        if dosya is not None:
            self.tutulan_dosya = dosya
            uzanti = self.list_donusum_str(self.uzanti_bul(dosya))
            self.tutulan_uzanti = uzanti
            if uzanti == ".db" or uzanti == ".sqlite":
                self.veri_yukle_sql(dosya)
            elif uzanti == ".csv":
                self.veri_yukle_csv(dosya)
            elif uzanti == ".xlsx":
                self.veri_yukle_xlsx(dosya)
            else:
                self.statusbar.showMessage("Bu tur dosyalar acilamaz")
        #except sqlite3.DatabaseError :
            #bildirim.uyari_kutusu("Yüklemeye çalıştığınız dosya bir veritabanı değildir.")
        #except Exception as hata:
            #self.statusbar.showMessage("Lutfen uygulamayi yeniden acin!Bir yukleme hatasi ile karsilasildi:"+str(hata),5000)
      
    def veri_isle_degistir(self,veriler):
        
        for asil, duzenlenmis in zip(self.tablo_, veriler):
            if asil != duzenlenmis:
                for index, (i, j) in enumerate(zip(asil, duzenlenmis)):
                    if i != j:
                        sorgu = f"UPDATE {self.tablo_adi[0][0]} SET " \
                            f"{self.columns[index]} = ? " \
                            f"WHERE {self.columns[1]} = ?"
                        self.imlec.execute(sorgu, (j, asil[1]))
        self.baglanti.commit()
        

    def veri_cek(self):
        
        veriler = []
        for satir in range(len(self.tablo_)):
            veri_satir = []
            for sutun in range(self.get_sutun_sayisi()):
                parcacik = self.tablo.item(satir, sutun)
                veri_satir.append(parcacik.text())
            veriler.append(veri_satir)
        self.veri_isle_degistir(veriler)
        
    
    def degistir(self):
        bildirim.mesaj_kutusu("DEGİSİM","Veri degisimi yapmak istediginize emin misiniz?")
        if self.kutu.clickedButton() == self.buttonY:
            self.veri_cek()
            self.temizle_tablo()
            self.guncel_tablo()
        elif self.kutu.calickedButton() == self.buttonN:
            bildirim.uyari_kutusu("Degisiminiz kaydedilmedi.")
            self.temizle_tablo()
            self.guncel_tablo()
    
    def get_sutun_sayisi(self):
        sutun_sorgusu = "PRAGMA table_info(%s)" % self.tablo_adi[0][0]
        self.imlec.execute(sutun_sorgusu)
        return len(self.imlec.fetchall())
    
    def veri_ekle(self,sm):
        try:
            self.imlec.execute(f"INSERT INTO {self.tablo_adi[0][0]} VALUES({', '.join('?' * self.get_sutun_sayisi())})",sm)
            self.baglanti.commit()
            self.temizle_tablo()
            self.guncel_tablo()
        except Exception as hata:
            self.statusbar.showMessage("Veri eklenemedi.Lutfen tekrar deneyin!"+str(hata),5000)
    
    def satir_ekle(self):
        
        bildirim.mesaj_kutusu("YENI KAYIT","Yeni bir kayit yapmak istediginize emin misiniz?(Ilerlediginizde hayvan listenizde degisim yapilacaktir.)")
        if self.kutu.clickedButton() == self.buttonY:
            bildirim.uyari_kutusu("Lutfen veri girisi yapmadan yeni mesajin gelmesini bekleyin!")
            gecici_satir_verisi = []
            for x in range(self.get_sutun_sayisi()):
                gecici_satir_verisi.append("")
            self.veri_ekle(gecici_satir_verisi)
            bildirim.uyari_kutusu("Yeni satiriniza veri girisi yapabilirsiniz.Bu asamadan sonra yeni verilerinizi 'DEGISTIR' butonu ile kaydedebilirsiniz.")
        else :
            bildirim.uyari_kutusu("Yeni kayit ekleme isleminiz iptal edildi!")
        
            #self.statusbar.showMessage("Malesef ekleme islemi basarisiz oldu.Lutfen tekrar deneyin!"+str(hata),5000)
    
    def silme(self):
        try:
            silinecek_satir = self.tablo.selectedItems()
            secilen = silinecek_satir[0].text()
            sorgu_silme = ("DELETE FROM {} WHERE {} = ?".format((self.tablo_adi[0][0]),(self.columns[0])))
            self.imlec.execute(sorgu_silme,(secilen,))
            self.baglanti.commit()
            self.temizle_tablo()
            self.guncel_tablo()
        except Exception as hata:
           self.statusbar.showMessage("Malesef silme islemi basarisiz oldu.Lutfen tekrar deneyin!"+str(hata),5000)
    
    def baglantilar(self):
        try:
            self.buton_yukle.clicked.connect(self.veri_yukle)
            self.buton_cik.clicked.connect(self.cikis)
            self.buton_degistir.clicked.connect(self.degistir)
            self.buton_ekle.clicked.connect(self.satir_ekle)
            self.buton_sil.clicked.connect(self.silme)
        except Exception as hata:
           self.statusbar.showMessage("Butonlarin baglantisi ile ilgili bir hata olustu.Lutfen uygulamayı yeniden baslatin!Hata:"+str(hata),5000)

    def ac_pencere(self):
        self.setupUi(self)
        self.baglantilar()
        